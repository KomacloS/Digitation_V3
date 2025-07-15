# ─── component_placer/bom_handler/bom_handler.py ───────────────────────────
from io import StringIO
from utils.file_ops import safe_write, rotate_backups   # ← NEW
import csv
import os
from typing import Dict, Any, List, Set
from logs.log_handler import LogHandler

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PyQt5.QtWidgets import QMessageBox

class BOMHandler:
    """
    Handles the Bill-of-Materials (BOM) for the PCB digitization project.
    Each entry in the BOM corresponds to a component (e.g., R1, R2, U4, etc.)
    and stores associated attributes: function, value, package, and part number.
    """

    def __init__(self):
        # Internal dictionary mapping component names to their attributes.
        # Example: "R1": {"function": "RESISTOR", "value": "10k", "package": "0805", "part_number": "XYZ123"}
        self.bom: Dict[str, Dict[str, str]] = {}
        self.log = LogHandler()
        self.log.info("BOMHandler initialized.", module="BOMHandler", func="__init__")

        # 1) Provide check_mismatch(...) exactly as you have
    def check_mismatch(self, board_components: List[str]) -> (Set[str], Set[str]):
        """Returns (missing, extra)."""
        # Missing = On board, not in BOM
        # Extra   = In BOM, not on board
        board_set = set(board_components)
        bom_set   = set(self.bom.keys())
        missing = board_set - bom_set
        extra   = bom_set - board_set
        return (missing, extra)


    def check_and_fix_mismatch(self, board_components: List[str], main_window, official_csv_path: str):
        """
        Given a list of board component names, compares these to the current BOM.
        Extra components (in the BOM but not on the board) are removed immediately
        and the updated BOM is saved to the official CSV path.
        If any missing components (on the board but not in the BOM) remain,
        opens the BOMEditorDialog so the user can add the missing entries.
        """
        missing, extra = self.check_mismatch(board_components)

        # Auto-remove extra components without asking the user.
        if extra:
            for comp in extra:
                self.log.log("info", f"Auto-removing extra component '{comp}' from BOM.",
                             module="BOMHandler", func="check_and_fix_mismatch")
                self.remove_component(comp)
            # Recalculate mismatch after auto-removal.
            missing, extra = self.check_mismatch(board_components)
            # Save the updated BOM immediately.
            if official_csv_path:
                if self.save_bom(official_csv_path):
                    self.log.log("info", f"BOM auto-saved after extra components removal to {official_csv_path}.",
                                 module="BOMHandler", func="check_and_fix_mismatch")
                else:
                    self.log.log("warning", "Failed to auto-save BOM after extra removal.",
                                 module="BOMHandler", func="check_and_fix_mismatch")

        if missing:
            self.log.log("info", "Missing components found => launching BOMEditorDialog for missing entries.",
                         module="BOMHandler", func="check_and_fix_mismatch")
            from component_placer.bom_handler.bom_editor_dialog import BOMEditorDialog
            board_set = set(board_components)
            dialog = BOMEditorDialog(
                bom_handler=self,
                board_component_names=board_set,
                parent=main_window
            )
            if dialog.exec_() == dialog.Accepted:
                if official_csv_path:
                    if self.save_bom(official_csv_path):
                        self.log.log("info", f"BOM updated and saved to {official_csv_path}.",
                                     module="BOMHandler", func="check_and_fix_mismatch")
                    else:
                        self.log.log("warning", "Failed to save updated BOM after mismatch fix.",
                                     module="BOMHandler", func="check_and_fix_mismatch")
                else:
                    self.log.log("warning", "No valid CSV path provided; cannot overwrite BOM on disk.",
                                 module="BOMHandler", func="check_and_fix_mismatch")
            else:
                self.log.log("info", "User canceled missing BOM editing. BOM remains with auto-removed extras.",
                             module="BOMHandler", func="check_and_fix_mismatch")
        else:
            self.log.log("info", "No missing components in BOM mismatch; extra components auto-removed.",
                         module="BOMHandler", func="check_and_fix_mismatch")

    def add_component(self, component_name: str, function: str, value: str, package: str, part_number: str) -> None:
        """
        Adds or updates a component entry in the BOM.
        """
        self.bom[component_name] = {
            "function": function,
            "value": value,
            "package": package,
            "part_number": part_number
        }

    def update_component(self, component_name: str, **kwargs) -> None:
        """
        Updates attributes of an existing component.
        Accepts keyword arguments for any of the keys: function, value, package, part_number.
        """
        if component_name in self.bom:
            for key, val in kwargs.items():
                if key in self.bom[component_name]:
                    self.bom[component_name][key] = val
            self.log.info(f"Updated component '{component_name}' with {kwargs}.",
                          module="BOMHandler", func="update_component")
            self.log.debug(f"Current BOM state: {self.bom}",
                           module="BOMHandler", func="update_component")
        else:
            self.log.warning(f"Attempted to update non-existent component '{component_name}'.",
                             module="BOMHandler", func="update_component")

    def remove_component(self, component_name: str) -> None:
        """
        Removes a component from the BOM.
        """
        if component_name in self.bom:
            del self.bom[component_name]
            self.log.info(f"Removed component '{component_name}' from BOM.",
                          module="BOMHandler", func="remove_component")
            self.log.debug(f"Current BOM state: {self.bom}",
                           module="BOMHandler", func="remove_component")
        else:
            self.log.warning(f"Attempted to remove non-existent component '{component_name}'.",
                             module="BOMHandler", func="remove_component")

    def get_component(self, component_name: str) -> Dict[str, str]:
        """
        Retrieves the attribute dictionary for a given component name.
        Returns an empty dictionary if the component is not found.
        """
        return self.bom.get(component_name, {})

    def get_all_components(self) -> List[Dict[str, Any]]:
        """
        Returns a list of all components as dictionaries, each including the component name.
        """
        result = []
        for comp_name, attrs in self.bom.items():
            comp_dict = {"component_name": comp_name}
            comp_dict.update(attrs)
            result.append(comp_dict)
        return result

    def save_bom(self, file_path: str, fixed_ts: str | None = None) -> bool:
        """
        Atomically saves the current BOM to *file_path* (CSV) and
        rotates backups: <central_backup_dir>/<project>/<file>.YYYYmmdd_HHMMSS.bak
        Returns True on success, False otherwise.
        """
        try:
            # 1) build CSV entirely in memory
            buf = StringIO()
            fieldnames = ["component_name", "function",
                          "value", "package", "part_number"]
            writer = csv.DictWriter(buf, fieldnames=fieldnames,
                                    lineterminator="\n")
            writer.writeheader()
            for component in self.get_all_components():
                writer.writerow(component)
            payload = buf.getvalue()

            # 2) create/rotate backup of current file (if it exists)
            rotate_backups(file_path, fixed_ts=fixed_ts)

            # 3) atomic write; old file is replaced only after fsync
            if safe_write(file_path, payload, encoding="utf-8"):
                self.log.info(f"BOM saved safely to '{file_path}'.",
                              module="BOMHandler", func="save_bom")
                return True
            else:
                self.log.error(f"safe_write failed for '{file_path}'.",
                               module="BOMHandler", func="save_bom")
                return False

        except Exception as e:
            self.log.error(f"Error saving BOM to '{file_path}': {e}",
                           module="BOMHandler", func="save_bom")
            return False
        
    def load_bom(self, file_path: str) -> bool:
        """
        Loads the BOM from a CSV file at the specified file_path.
        The CSV file is expected to have a header with the columns:
        component_name, function, value, package, part_number.

        Returns True if loading was successful, False otherwise.
        """
        if not os.path.exists(file_path):
            self.log.warning(f"BOM file '{file_path}' does not exist.",
                             module="BOMHandler", func="load_bom")
            return False

        try:
            with open(file_path, "r", newline="") as csvfile:
                reader = csv.DictReader(csvfile)
                self.bom = {}  # Clear existing BOM.
                for row in reader:
                    comp_name = row.get("component_name", "").strip()
                    if comp_name:
                        self.bom[comp_name] = {
                            "function": row.get("function", "").strip(),
                            "value": row.get("value", "").strip(),
                            "package": row.get("package", "").strip(),
                            "part_number": row.get("part_number", "").strip()
                        }
            self.log.info(f"BOM loaded successfully from '{file_path}'.",
                          module="BOMHandler", func="load_bom")
            self.log.debug(f"Loaded BOM state: {self.bom}",
                           module="BOMHandler", func="load_bom")
            return True
        except Exception as e:
            self.log.error(f"Error loading BOM from '{file_path}': {e}",
                           module="BOMHandler", func="load_bom")
            return False

    # --------------------------------------------------------------------------
    #                           Mismatch Checking
    # --------------------------------------------------------------------------
    def check_mismatch(self, board_component_names: List[str]) -> (Set[str], Set[str]):
        """
        Given a list of component names from the board (ObjectLibrary),
        returns a tuple: (missing, extra)
          - missing: components on the board that aren't in the BOM
          - extra: components in the BOM that aren't on the board
        """
        board_set = set(board_component_names)
        bom_set = set(self.bom.keys())

        missing = board_set - bom_set  # On board, not in BOM
        extra   = bom_set - board_set  # In BOM, not on board

        return (missing, extra)

    def handle_mismatch(self, missing: Set[str], extra: Set[str], parent_widget, bom_path: str) -> None:
        """
        If missing or extra is non-empty, show a pop-up with 'Edit' or 'Cancel'.
        If user clicks 'Edit', generate a color-coded .xlsx with mismatches.
        The file is then opened in Excel (Windows only).
        """
        if not missing and not extra:
            return  # No mismatch => do nothing

        msg = []
        if missing:
            msg.append(f"Missing in BOM: {', '.join(sorted(missing))}")
        if extra:
            msg.append(f"Extra in BOM: {', '.join(sorted(extra))}")
        msg_text = "\n".join(msg)
        msg_text += "\n\nChoose 'Edit' to fix in Excel or 'Cancel' to ignore."

        box = QMessageBox(parent_widget)
        box.setWindowTitle("BOM Mismatch Detected")
        box.setText(msg_text)
        edit_btn = box.addButton("Edit", QMessageBox.AcceptRole)
        cancel_btn = box.addButton("Cancel", QMessageBox.RejectRole)
        box.setDefaultButton(edit_btn)
        box.exec_()
        clicked = box.clickedButton()
        if clicked == edit_btn:
            self.generate_mismatch_spreadsheet(bom_path, missing, extra)
        else:
            # Cancel => do nothing
            pass

    def generate_mismatch_spreadsheet(self, bom_path: str, missing: Set[str], extra: Set[str]) -> None:
        """
        Creates a color-coded .xlsx file indicating extra or missing components.
        Opens it in Excel (Windows).
        If openpyxl is not available, logs a warning instead.
        """
        if not OPENPYXL_AVAILABLE:
            self.log.warning("openpyxl is not installed; cannot generate mismatch spreadsheet.")
            return

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        import os

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM Mismatch"

        headers = ["component_name", "function", "value", "package", "part_number", "Status"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx).value = header

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Dump the existing BOM components
        row = 2
        for comp_name, attrs in self.bom.items():
            ws.cell(row=row, column=1).value = comp_name
            ws.cell(row=row, column=2).value = attrs.get("function", "")
            ws.cell(row=row, column=3).value = attrs.get("value", "")
            ws.cell(row=row, column=4).value = attrs.get("package", "")
            ws.cell(row=row, column=5).value = attrs.get("part_number", "")

            if comp_name in extra:
                ws.cell(row=row, column=6).value = "EXTRA"
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = red_fill
            else:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = normal_fill
            row += 1

        # Add rows for missing components
        for comp in missing:
            ws.cell(row=row, column=1).value = comp
            ws.cell(row=row, column=6).value = "MISSING"
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = red_fill
            row += 1

        # Save as .xlsx next to the CSV (or you can choose another path)
        mismatch_xlsx = os.path.splitext(bom_path)[0] + "_mismatch.xlsx"
        wb.save(mismatch_xlsx)
        self.log.info(f"Mismatch spreadsheet created: {mismatch_xlsx}",
                      module="BOMHandler", func="generate_mismatch_spreadsheet")

        # Attempt to open in Excel on Windows
        try:
            os.startfile(mismatch_xlsx)
        except Exception as e:
            self.log.warning(f"Could not open Excel file '{mismatch_xlsx}': {e}",
                             module="BOMHandler", func="generate_mismatch_spreadsheet")


    def import_from_mismatch_xlsx(self, xlsx_path: str) -> bool:
        """
        Reads the updated mismatch XLSX and merges it into the BOM:
        - Clears the existing BOM in memory
        - Rebuilds from the rows that remain
        - Ignores any row that lacks a 'component_name'
        - Ignores the 'status' column if present
        - If user removed a row entirely, that component is removed
        - If user changed the name, it becomes a new entry
        Returns True if successful, False otherwise.
        """
        try:
            import openpyxl
        except ImportError:
            self.log.warning("openpyxl not installed; cannot import XLSX.")
            return False

        if not os.path.exists(xlsx_path):
            self.log.warning(f"'{xlsx_path}' not found; cannot import mismatch XLSX.")
            return False

        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active  # assume the first sheet is what we want

        # Optional: you might check the header row to confirm correct columns
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            header_val = str(cell.value).strip().lower() if cell.value else ""
            headers[header_val] = col_idx

        required_columns = ["component_name", "function", "value", "package", "part_number"]
        for col in required_columns:
            if col not in headers:
                self.log.warning(f"Mismatch XLSX missing required column '{col}'")
                return False

        # Clear current BOM and rebuild from the spreadsheet rows
        new_bom = {}
        row_count = ws.max_row
        for row_idx in range(2, row_count+1):
            comp_name = ws.cell(row=row_idx, column=headers["component_name"]).value
            if not comp_name:
                continue  # skip any row missing a component_name
            comp_name = str(comp_name).strip()

            # Gather function, value, package, part_number
            function    = str(ws.cell(row=row_idx, column=headers["function"]).value or "").strip()
            value       = str(ws.cell(row=row_idx, column=headers["value"]).value or "").strip()
            package     = str(ws.cell(row=row_idx, column=headers["package"]).value or "").strip()
            part_number = str(ws.cell(row=row_idx, column=headers["part_number"]).value or "").strip()

            # Insert into new_bom
            new_bom[comp_name] = {
                "function": function,
                "value": value,
                "package": package,
                "part_number": part_number
            }

        self.bom = new_bom
        self.log.info(f"Rebuilt BOM from '{xlsx_path}' with {len(new_bom)} entries.",
                    module="BOMHandler", func="import_from_mismatch_xlsx")
        self.log.debug(f"New BOM state: {self.bom}", module="BOMHandler", func="import_from_mismatch_xlsx")
        return True
