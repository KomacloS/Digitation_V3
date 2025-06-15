# component_placer/bom_handler.py

import csv
import os
import copy
from typing import Dict, Any, List, Set
from logs.log_handler import LogHandler

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PyQt5.QtWidgets import QMessageBox

class BOMHandler:
    """
    Handles the Bill-of-Materials (BOM) for the PCB digitization project.
    Each entry in the BOM corresponds to a component (e.g., R1, R2, U4, etc.)
    and stores associated attributes: function, value, package, and part number.
    """

    def __init__(self):
        # Internal dictionary mapping component names to their attributes.
        # Example: "R1": {"function": "RESISTOR", "value": "10k", "package": "0805", "part_number": "XYZ123"}
        self.bom: Dict[str, Dict[str, str]] = {}
        self.log = LogHandler()
        self.log.info("BOMHandler initialized.", module="BOMHandler", func="__init__")

    def add_component(self, component_name: str, function: str, value: str, package: str, part_number: str) -> None:
        self.bom[component_name] = {
            "function": function,
            "value": value,
            "package": package,
            "part_number": part_number
        }
        self.log.info(f"Added/updated component '{component_name}' to BOM.",
                      module="BOMHandler", func="add_component")
        self.log.debug(f"Current BOM state after addition: {self.bom}",
                       module="BOMHandler", func="add_component")

    def update_component(self, component_name: str, **kwargs) -> None:
        """
        Updates attributes of an existing component.
        Accepts keyword arguments for any of the keys: function, value, package, part_number.
        """
        if component_name in self.bom:
            for key, val in kwargs.items():
                if key in self.bom[component_name]:
                    self.bom[component_name][key] = val
            self.log.info(f"Updated component '{component_name}' with {kwargs}.",
                          module="BOMHandler", func="update_component")
            self.log.debug(f"Current BOM state: {self.bom}",
                           module="BOMHandler", func="update_component")
        else:
            self.log.warning(f"Attempted to update non-existent component '{component_name}'.",
                             module="BOMHandler", func="update_component")

    def remove_component(self, component_name: str) -> None:
        """
        Removes a component from the BOM.
        """
        if component_name in self.bom:
            del self.bom[component_name]
            self.log.info(f"Removed component '{component_name}' from BOM.",
                          module="BOMHandler", func="remove_component")
            self.log.debug(f"Current BOM state: {self.bom}",
                           module="BOMHandler", func="remove_component")
        else:
            self.log.warning(f"Attempted to remove non-existent component '{component_name}'.",
                             module="BOMHandler", func="remove_component")

    def get_component(self, component_name: str) -> Dict[str, str]:
        """
        Retrieves the attribute dictionary for a given component name.
        Returns an empty dictionary if the component is not found.
        """
        return self.bom.get(component_name, {})

    def get_all_components(self) -> List[Dict[str, Any]]:
        """
        Returns a list of all components as dictionaries, each including the component name.
        """
        result = []
        for comp_name, attrs in self.bom.items():
            comp_dict = {"component_name": comp_name}
            comp_dict.update(attrs)
            result.append(comp_dict)
        return result

    def save_bom(self, file_path: str) -> bool:
        """
        Saves the current BOM to a CSV file at the specified file_path.
        The CSV file will have a header with the columns:
        component_name, function, value, package, part_number.

        Returns True if saving was successful, False otherwise.
        """
        try:
            # If the file already exists, try to remove it.
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    self.log.info(f"Existing BOM file '{file_path}' removed.",
                                  module="BOMHandler", func="save_bom")
                except Exception as e:
                    self.log.error(f"Cannot remove existing BOM file '{file_path}': {e}",
                                   module="BOMHandler", func="save_bom")
                    return False

            with open(file_path, "w", newline="") as csvfile:
                fieldnames = ["component_name", "function", "value", "package", "part_number"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                for component in self.get_all_components():
                    writer.writerow(component)
            self.log.info(f"BOM saved successfully to '{file_path}'.",
                          module="BOMHandler", func="save_bom")
            return True
        except Exception as e:
            self.log.error(f"Error saving BOM to '{file_path}': {e}",
                           module="BOMHandler", func="save_bom")
            return False

    def load_bom(self, file_path: str) -> bool:
        """
        Loads the BOM from a CSV file at the specified file_path.
        The CSV file is expected to have a header with the columns:
        component_name, function, value, package, part_number.

        Returns True if loading was successful, False otherwise.
        """
        if not os.path.exists(file_path):
            self.log.warning(f"BOM file '{file_path}' does not exist.",
                             module="BOMHandler", func="load_bom")
            return False

        try:
            with open(file_path, "r", newline="") as csvfile:
                reader = csv.DictReader(csvfile)
                self.bom = {}  # Clear existing BOM.
                for row in reader:
                    comp_name = row.get("component_name", "").strip()
                    if comp_name:
                        self.bom[comp_name] = {
                            "function": row.get("function", "").strip(),
                            "value": row.get("value", "").strip(),
                            "package": row.get("package", "").strip(),
                            "part_number": row.get("part_number", "").strip()
                        }
            self.log.info(f"BOM loaded successfully from '{file_path}'.",
                          module="BOMHandler", func="load_bom")
            self.log.debug(f"Loaded BOM state: {self.bom}",
                           module="BOMHandler", func="load_bom")
            return True
        except Exception as e:
            self.log.error(f"Error loading BOM from '{file_path}': {e}",
                           module="BOMHandler", func="load_bom")
            return False

    # --------------------------------------------------------------------------
    #                           Mismatch Checking
    # --------------------------------------------------------------------------
    def check_mismatch(self, board_component_names: List[str]) -> (Set[str], Set[str]):
        """
        Given a list of component names from the board (ObjectLibrary),
        returns a tuple: (missing, extra)
          - missing: components on the board that aren't in the BOM
          - extra: components in the BOM that aren't on the board
        """
        board_set = set(board_component_names)
        bom_set = set(self.bom.keys())

        missing = board_set - bom_set  # On board, not in BOM
        extra   = bom_set - board_set  # In BOM, not on board

        return (missing, extra)

    def handle_mismatch(self, missing: Set[str], extra: Set[str], parent_widget, bom_path: str) -> None:
        """
        If missing or extra is non-empty, show a pop-up with 'Edit' or 'Cancel'.
        If user clicks 'Edit', generate a color-coded .xlsx with mismatches.
        The file is then opened in Excel (Windows only).
        """
        if not missing and not extra:
            return  # No mismatch => do nothing

        msg = []
        if missing:
            msg.append(f"Missing in BOM: {', '.join(sorted(missing))}")
        if extra:
            msg.append(f"Extra in BOM: {', '.join(sorted(extra))}")
        msg_text = "\n".join(msg)
        msg_text += "\n\nChoose 'Edit' to fix in Excel or 'Cancel' to ignore."

        box = QMessageBox(parent_widget)
        box.setWindowTitle("BOM Mismatch Detected")
        box.setText(msg_text)
        edit_btn = box.addButton("Edit", QMessageBox.AcceptRole)
        cancel_btn = box.addButton("Cancel", QMessageBox.RejectRole)
        box.setDefaultButton(edit_btn)
        box.exec_()
        clicked = box.clickedButton()
        if clicked == edit_btn:
            self.generate_mismatch_spreadsheet(bom_path, missing, extra)
        else:
            # Cancel => do nothing
            pass

    def generate_mismatch_spreadsheet(self, bom_path: str, missing: Set[str], extra: Set[str]) -> None:
        """
        Creates a color-coded .xlsx file indicating extra or missing components.
        Opens it in Excel (Windows).
        If openpyxl is not available, logs a warning instead.
        """
        if not OPENPYXL_AVAILABLE:
            self.log.warning("openpyxl is not installed; cannot generate mismatch spreadsheet.")
            return

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        import os

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM Mismatch"

        headers = ["component_name", "function", "value", "package", "part_number", "Status"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx).value = header

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Dump the existing BOM components
        row = 2
        for comp_name, attrs in self.bom.items():
            ws.cell(row=row, column=1).value = comp_name
            ws.cell(row=row, column=2).value = attrs.get("function", "")
            ws.cell(row=row, column=3).value = attrs.get("value", "")
            ws.cell(row=row, column=4).value = attrs.get("package", "")
            ws.cell(row=row, column=5).value = attrs.get("part_number", "")

            if comp_name in extra:
                ws.cell(row=row, column=6).value = "EXTRA"
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = red_fill
            else:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = normal_fill
            row += 1

        # Add rows for missing components
        for comp in missing:
            ws.cell(row=row, column=1).value = comp
            ws.cell(row=row, column=6).value = "MISSING"
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = red_fill
            row += 1

        # Save as .xlsx next to the CSV (or you can choose another path)
        mismatch_xlsx = os.path.splitext(bom_path)[0] + "_mismatch.xlsx"
        wb.save(mismatch_xlsx)
        self.log.info(f"Mismatch spreadsheet created: {mismatch_xlsx}",
                      module="BOMHandler", func="generate_mismatch_spreadsheet")

        # Attempt to open in Excel on Windows
        try:
            os.startfile(mismatch_xlsx)
        except Exception as e:
            self.log.warning(f"Could not open Excel file '{mismatch_xlsx}': {e}",
                             module="BOMHandler", func="generate_mismatch_spreadsheet")


    def push_state(self):
        """
        Pushes a deep copy of the current BOM state to the shared undo/redo manager.
        """
        if self.undo_redo_manager:
            state = {"bom": copy.deepcopy(self.bom)}
            self.undo_redo_manager.push_state(extra_state=state)
            self.log.log("debug", f"Pushed BOM state to undo/redo stack: {state}",
                         module="BOMHandler", func="push_state")
        else:
            self.log.log("warning", "Undo/Redo manager not set in BOMHandler; BOM state not saved.",
                         module="BOMHandler", func="push_state")